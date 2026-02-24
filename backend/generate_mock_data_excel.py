"""
UrbanFlow — Mock Data Excel Generator
Generates the exact mock data that the system produces when running
without SUMO, YOLO, or PyTorch (current operational mode).
"""

import random
import math
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT    = Font(name="Calibri", size=14, bold=True, color="0F172A")
SUB_FONT      = Font(name="Calibri", size=11, bold=True, color="6366F1")
NOTE_FONT     = Font(name="Calibri", size=10, italic=True, color="64748B")
DATA_FONT     = Font(name="Calibri", size=10)
GREEN_FILL    = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
YELLOW_FILL   = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
RED_FILL      = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
MOCK_FILL     = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
ALT_FILL      = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
BORDER        = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)
phase_fill = {"green": GREEN_FILL, "yellow": YELLOW_FILL, "red": RED_FILL,
              "GREEN": GREEN_FILL, "YELLOW": YELLOW_FILL, "RED": RED_FILL}


def header_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER


def style_rows(ws, r1, r2, cols):
    for r in range(r1, r2 + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER
            if (r - r1) % 2 == 1:
                cell.fill = ALT_FILL


def auto_col(ws, cols):
    for c in range(1, cols + 1):
        mx = max((len(str(ws.cell(r, c).value or "")) for r in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(c)].width = max(mx + 3, 12)


random.seed(42)  # Reproducible mock data
base = datetime(2026, 2, 20, 22, 30, 0)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Camera Server Mock Frames (handle_frame mock path)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Camera Mock Frames"

ws1.merge_cells('A1:K1')
ws1.cell(row=1, column=1, value="Mock Data — Camera Server handle_frame() Output (CV_AVAILABLE=False)").font = TITLE_FONT
ws1.cell(row=2, column=1, value="Source: mobile_camera_server.py → handle_frame() mock path").font = NOTE_FONT

h1 = [
    "Timestamp", "Frame #", "Mock",
    "Signal State", "Signal Duration (s)", "Signal Confidence",
    "Total Vehicles", "Queue Length", "Avg Speed (m/s)",
    "RL Reward", "Data Source"
]
header_row(ws1, 4, h1)

row = 5
for f in range(1, 201):
    ts = base + timedelta(seconds=f * 0.2)
    state = random.choice(["red", "green", "yellow"])
    dur = round(random.uniform(5.0, 30.0), 1)
    conf = round(random.uniform(0.6, 0.95), 2)
    vehs = random.randint(3, 12)
    queue = random.randint(0, vehs)
    speed = round(random.uniform(5.0, 15.0), 1)
    reward = -float(queue)

    ws1.cell(row=row, column=1, value=ts.strftime("%Y-%m-%d %H:%M:%S.%f")[:-4])
    ws1.cell(row=row, column=2, value=f)
    ws1.cell(row=row, column=3, value="TRUE")
    ws1.cell(row=row, column=4, value=state.upper())
    ws1.cell(row=row, column=5, value=dur)
    ws1.cell(row=row, column=6, value=conf)
    ws1.cell(row=row, column=7, value=vehs)
    ws1.cell(row=row, column=8, value=queue)
    ws1.cell(row=row, column=9, value=speed)
    ws1.cell(row=row, column=10, value=reward)
    ws1.cell(row=row, column=11, value="MOCK (random)")
    ws1.cell(row=row, column=3).fill = MOCK_FILL
    if state in phase_fill:
        ws1.cell(row=row, column=4).fill = phase_fill[state]
    row += 1

style_rows(ws1, 5, row - 1, len(h1))
auto_col(ws1, len(h1))

# Chart: Vehicle count over frames
ch1 = LineChart()
ch1.title = "Mock — Vehicle Count per Frame"
ch1.x_axis.title = "Frame"
ch1.y_axis.title = "Vehicles"
ch1.style = 10
ch1.width = 28
ch1.height = 14
d1 = Reference(ws1, min_col=7, min_row=4, max_row=row - 1)
c1 = Reference(ws1, min_col=2, min_row=5, max_row=row - 1)
ch1.add_data(d1, titles_from_data=True)
ch1.set_categories(c1)
ch1.series[0].graphicalProperties.line.width = 15000
ws1.add_chart(ch1, "A" + str(row + 2))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Simulation Mock Data (SUMO mock_mode=True)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Simulation Mock Data")

ws2.merge_cells('A1:M1')
ws2.cell(row=1, column=1, value="Mock Data — SUMO Environment (mock_mode=True, TRACI_AVAILABLE=False)").font = TITLE_FONT
ws2.cell(row=2, column=1, value="Source: environment.py → SumoEnvironment.get_state() mock path").font = NOTE_FONT

h2 = [
    "Timestamp", "Step #", "Intersection ID",
    "Queue Length", "Vehicle Count", "Avg Speed (m/s)",
    "Phase (0=G 1=Y 2=R)", "Phase Name",
    "RL Action", "RL Reward",
    "Waiting Time (est.)", "Stopped Vehicles (est.)", "Mode"
]
header_row(ws2, 4, h2)

row = 5
ints = ["INT_1", "INT_2", "INT_3", "INT_4"]
for step in range(1, 121):
    ts = base + timedelta(seconds=step)
    for iid in ints:
        q = round(random.uniform(0, 20), 1)
        v = round(random.uniform(5, 50), 1)
        s = round(random.uniform(5, 15), 1)
        p = random.choice([0, 1, 2])
        pname = ["GREEN", "YELLOW", "RED"][p]
        action = random.randint(0, 3)
        reward = round(random.uniform(-100, -10), 1)
        wait = round(q * random.uniform(2, 4), 1)
        stopped = int(q * random.uniform(0.3, 0.7))

        ws2.cell(row=row, column=1, value=ts.strftime("%H:%M:%S"))
        ws2.cell(row=row, column=2, value=step)
        ws2.cell(row=row, column=3, value=iid)
        ws2.cell(row=row, column=4, value=q)
        ws2.cell(row=row, column=5, value=v)
        ws2.cell(row=row, column=6, value=s)
        ws2.cell(row=row, column=7, value=p)
        ws2.cell(row=row, column=8, value=pname)
        ws2.cell(row=row, column=9, value=action)
        ws2.cell(row=row, column=10, value=reward)
        ws2.cell(row=row, column=11, value=wait)
        ws2.cell(row=row, column=12, value=stopped)
        ws2.cell(row=row, column=13, value="MOCK")

        ws2.cell(row=row, column=8).fill = phase_fill.get(pname, MOCK_FILL)
        ws2.cell(row=row, column=13).fill = MOCK_FILL
        row += 1

style_rows(ws2, 5, row - 1, len(h2))
auto_col(ws2, len(h2))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: RL Agent Mock Actions (TORCH_AVAILABLE=False)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("RL Agent Mock Actions")

ws3.merge_cells('A1:I1')
ws3.cell(row=1, column=1, value="Mock Data — RL Agent (TORCH_AVAILABLE=False, Dummy Random Actions)").font = TITLE_FONT
ws3.cell(row=2, column=1, value="Source: rl_agent.py → RLAgent.get_action() dummy path").font = NOTE_FONT

h3 = [
    "Step", "Intersection",
    "State Vector [q, v, s, p]",
    "Action (random 0-3)", "Action Name",
    "Log Prob", "Value Estimate",
    "Reward", "Mode"
]
header_row(ws3, 4, h3)

action_names = {0: "GREEN", 1: "YELLOW", 2: "RED", 3: "EXTENDED GREEN"}
row = 5
for step in range(1, 201):
    iid = ints[(step - 1) % 4]
    q = round(random.uniform(0, 20), 1)
    v = round(random.uniform(5, 50), 1)
    s = round(random.uniform(5, 15), 1)
    p = random.choice([0, 1, 2])
    state_str = f"[{q}, {v}, {s}, {p}]"

    action = random.randint(0, 3)
    reward = round(random.uniform(-100, -10), 1)

    ws3.cell(row=row, column=1, value=step)
    ws3.cell(row=row, column=2, value=iid)
    ws3.cell(row=row, column=3, value=state_str)
    ws3.cell(row=row, column=4, value=action)
    ws3.cell(row=row, column=5, value=action_names[action])
    ws3.cell(row=row, column=6, value=0.0)    # log_prob = 0 in dummy mode
    ws3.cell(row=row, column=7, value=0.0)    # value = 0 in dummy mode
    ws3.cell(row=row, column=8, value=reward)
    ws3.cell(row=row, column=9, value="DUMMY")

    ws3.cell(row=row, column=5).fill = phase_fill.get(action_names[action], MOCK_FILL)
    ws3.cell(row=row, column=9).fill = MOCK_FILL
    row += 1

style_rows(ws3, 5, row - 1, len(h3))
auto_col(ws3, len(h3))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Dashboard WebSocket Broadcast Mock
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Dashboard Broadcast")

ws4.merge_cells('A1:K1')
ws4.cell(row=1, column=1, value="Mock Data — WebSocket Broadcast to Dashboard (traffic_update events)").font = TITLE_FONT
ws4.cell(row=2, column=1, value="Source: simulation.py → broadcast_loop() every 2 seconds").font = NOTE_FONT

h4 = [
    "Broadcast #", "Timestamp (ms)",
    "INT_1 Phase", "INT_1 Vehicles", "INT_1 Wait (s)",
    "INT_2 Phase", "INT_2 Vehicles", "INT_2 Wait (s)",
    "INT_3 Phase", "INT_3 Vehicles", "INT_3 Wait (s)",
]
header_row(ws4, 4, h4)

row = 5
for b in range(1, 151):
    ts_ms = int((base + timedelta(seconds=b * 2)).timestamp() * 1000)

    ws4.cell(row=row, column=1, value=b)
    ws4.cell(row=row, column=2, value=ts_ms)

    for i in range(3):
        col_base = 3 + i * 3
        ph = random.choice(["green", "yellow", "red"])
        vehs = int(random.uniform(5, 50))
        wait = round(random.uniform(0, 20), 1)

        ws4.cell(row=row, column=col_base, value=ph.upper())
        ws4.cell(row=row, column=col_base + 1, value=vehs)
        ws4.cell(row=row, column=col_base + 2, value=wait)
        ws4.cell(row=row, column=col_base).fill = phase_fill.get(ph, MOCK_FILL)

    row += 1

style_rows(ws4, 5, row - 1, len(h4))
auto_col(ws4, len(h4))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Data Collection Mock Samples
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Collected Samples")

ws5.merge_cells('A1:I1')
ws5.cell(row=1, column=1, value="Mock Data — Data Collection Samples (collected_data[])").font = TITLE_FONT
ws5.cell(row=2, column=1, value="Source: handle_frame() → is_collecting=True → collected_data.append()").font = NOTE_FONT

h5 = [
    "Sample #", "Timestamp",
    "State[0] (green?)", "State[1] (yellow?)", "State[2] (red?)", "State[3] (queue)",
    "Reward", "Vehicles", "Source"
]
header_row(ws5, 4, h5)

row = 5
for s in range(1, 101):
    ts = base + timedelta(seconds=s * 0.5)
    state = random.choice(["green", "yellow", "red"])
    q = random.randint(0, 12)

    ws5.cell(row=row, column=1, value=s)
    ws5.cell(row=row, column=2, value=ts.strftime("%Y-%m-%d %H:%M:%S"))
    ws5.cell(row=row, column=3, value=1.0 if state == "green" else 0.0)
    ws5.cell(row=row, column=4, value=1.0 if state == "yellow" else 0.0)
    ws5.cell(row=row, column=5, value=1.0 if state == "red" else 0.0)
    ws5.cell(row=row, column=6, value=float(q))
    ws5.cell(row=row, column=7, value=-float(q))
    ws5.cell(row=row, column=8, value=random.randint(3, 12))
    ws5.cell(row=row, column=9, value="MOCK")
    ws5.cell(row=row, column=9).fill = MOCK_FILL
    row += 1

style_rows(ws5, 5, row - 1, len(h5))
auto_col(ws5, len(h5))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: API Response Mock Snapshots
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("API Responses")

ws6.merge_cells('A1:E1')
ws6.cell(row=1, column=1, value="Mock Data — API Endpoint Response Snapshots").font = TITLE_FONT

# /api/status
ws6.cell(row=3, column=1, value="/api/status").font = SUB_FONT
h6a = ["Field", "Value", "Type", "Description"]
header_row(ws6, 4, h6a)

status_data = [
    ("status",          "running",  "string",  "Server health status"),
    ("cv_available",    "True",     "boolean", "OpenCV loaded successfully"),
    ("uptime",          "342.5",    "float",   "Seconds since server start"),
    ("frame_count",     "0",        "int",     "Total frames processed (0 = no phone)"),
    ("avg_fps",         "0.0",      "float",   "Average processing FPS"),
    ("connected_clients", "0",      "int",     "Active Socket.IO clients"),
]
for i, (f, v, t, d) in enumerate(status_data):
    r = 5 + i
    ws6.cell(row=r, column=1, value=f)
    ws6.cell(row=r, column=2, value=v)
    ws6.cell(row=r, column=3, value=t)
    ws6.cell(row=r, column=4, value=d)
style_rows(ws6, 5, 5 + len(status_data) - 1, 4)

# /api/rl_state
r_start = 5 + len(status_data) + 2
ws6.cell(row=r_start, column=1, value="/api/rl_state").font = SUB_FONT
header_row(ws6, r_start + 1, ["Field", "Value", "Type", "Description"])

rl_data = [
    ("state",     "[0.0, 0.0, ..., 0.0]", "float[32]", "32-dim state vector (all zeros when no data)"),
    ("reward",    "0.0",                    "float",     "Current reward (0.0 in mock mode)"),
    ("timestamp", "2026-02-20T22:30:00",   "string",    "ISO timestamp of state"),
]
for i, (f, v, t, d) in enumerate(rl_data):
    r = r_start + 2 + i
    ws6.cell(row=r, column=1, value=f)
    ws6.cell(row=r, column=2, value=v)
    ws6.cell(row=r, column=3, value=t)
    ws6.cell(row=r, column=4, value=d)
style_rows(ws6, r_start + 2, r_start + 2 + len(rl_data) - 1, 4)

# /api/debug (mock mode)
r_start2 = r_start + 2 + len(rl_data) + 2
ws6.cell(row=r_start2, column=1, value="/api/debug (mock mode)").font = SUB_FONT
header_row(ws6, r_start2 + 1, ["Field", "Value", "Type", "Description"])

debug_data = [
    ("status",       "mock",    "string",  "Running in mock mode"),
    ("cv_available", "False",   "boolean", "OpenCV not available"),
    ("frame_count",  "0",       "int",     "Frames processed"),
    ("message",      "Running in mock mode", "string", "Human-readable status"),
]
for i, (f, v, t, d) in enumerate(debug_data):
    r = r_start2 + 2 + i
    ws6.cell(row=r, column=1, value=f)
    ws6.cell(row=r, column=2, value=v)
    ws6.cell(row=r, column=3, value=t)
    ws6.cell(row=r, column=4, value=d)
style_rows(ws6, r_start2 + 2, r_start2 + 2 + len(debug_data) - 1, 4)

auto_col(ws6, 4)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7: Mock Data Ranges Reference
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Mock Data Ranges")

ws7.merge_cells('A1:F1')
ws7.cell(row=1, column=1, value="Mock Data — Value Ranges & Generation Logic").font = TITLE_FONT
ws7.cell(row=2, column=1, value="What random values does each mock path generate?").font = NOTE_FONT

h7 = ["Parameter", "Min", "Max", "Type", "Source Function", "Code Line"]
header_row(ws7, 4, h7)

ranges = [
    ("mock_state (signal)",    "red/green/yellow", "-",     "random.choice",   "handle_frame() mock path",      "~line 700"),
    ("mock_vehicles",          "3",     "12",    "random.randint",  "handle_frame() mock path",      "~line 701"),
    ("mock_speed",             "5.0",   "15.0",  "random.uniform",  "handle_frame() mock path",      "~line 702"),
    ("mock_queue",             "0",     "vehicles", "random.randint", "handle_frame() mock path",    "~line 703"),
    ("mock_duration",          "5.0",   "30.0",  "random.uniform",  "handle_frame() mock path",      "~line 704"),
    ("mock_confidence",        "0.6",   "0.95",  "random.uniform",  "handle_frame() mock path",      "~line 705"),
    ("sumo queue_length",      "0",     "20",    "random.uniform",  "SumoEnvironment.get_state()",   "~line 92"),
    ("sumo num_vehicles",      "5",     "50",    "random.uniform",  "SumoEnvironment.get_state()",   "~line 93"),
    ("sumo avg_speed",         "5",     "15",    "random.uniform",  "SumoEnvironment.get_state()",   "~line 94"),
    ("sumo phase",             "0",     "2",     "random.choice",   "SumoEnvironment.get_state()",   "~line 95"),
    ("sumo reward",            "-100",  "-10",   "random.uniform",  "SumoEnvironment.get_reward()",  "~line 137"),
    ("rl_agent action",        "0",     "3",     "random.randint",  "RLAgent.get_action() dummy",    "~line 80"),
    ("rl_agent log_prob",      "0.0",   "0.0",   "constant",        "RLAgent.get_action() dummy",    "~line 80"),
    ("rl_agent value",         "0.0",   "0.0",   "constant",        "RLAgent.get_action() dummy",    "~line 80"),
    ("motion fallback speed",  "5.0",   "15.0",  "np.random.uniform", "vehicle_detector.py detect()", "~line 125"),
]

row = 5
for param, mn, mx, typ, src, line in ranges:
    ws7.cell(row=row, column=1, value=param)
    ws7.cell(row=row, column=2, value=mn)
    ws7.cell(row=row, column=3, value=mx)
    ws7.cell(row=row, column=4, value=typ)
    ws7.cell(row=row, column=5, value=src)
    ws7.cell(row=row, column=6, value=line)
    row += 1

style_rows(ws7, 5, row - 1, len(h7))
auto_col(ws7, len(h7))


# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
path = "D:/Hackathon/UrbanFlow_Mock_Data.xlsx"
wb.save(path)
print("Mock data Excel saved: " + path)
print("Sheets: " + str(wb.sheetnames))
print("Done!")

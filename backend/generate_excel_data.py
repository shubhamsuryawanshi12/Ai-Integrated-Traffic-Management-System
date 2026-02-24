"""
UrbanFlow — Generate Excel Data File
Creates a comprehensive Excel workbook with sample/simulated data
that represents all the data the system works with.
"""

import random
import math
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color Palette ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT    = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
SUBTITLE_FONT = Font(name="Calibri", size=11, bold=True, color="3B82F6")
DATA_FONT     = Font(name="Calibri", size=10)
GREEN_FILL    = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
YELLOW_FILL   = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
RED_FILL      = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
ALT_ROW_FILL  = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
THIN_BORDER   = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)


def style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL


def auto_width(ws, col_count):
    for c in range(1, col_count + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c)].width = max(max_len + 3, 12)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Traffic Simulation Data (60 minutes, every 1 minute)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Traffic Simulation"

ws1.merge_cells('A1:L1')
ws1.cell(row=1, column=1).value = "UrbanFlow — Traffic Simulation Data (4 Intersections × 60 Steps)"
ws1.cell(row=1, column=1).font = TITLE_FONT

headers1 = [
    "Timestamp", "Step", "Intersection",
    "Queue Length", "Vehicle Count", "Avg Speed (m/s)",
    "Signal Phase", "Phase Duration (s)",
    "RL Action", "RL Reward",
    "Waiting Time (s)", "Stopped Vehicles"
]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=c, value=h)
style_header(ws1, 3, len(headers1))

base_time = datetime(2026, 2, 20, 8, 0, 0)  # 8 AM morning rush
intersections = ["INT_1", "INT_2", "INT_3", "INT_4"]
phases = ["GREEN", "YELLOW", "RED"]
phase_colors = {"GREEN": GREEN_FILL, "YELLOW": YELLOW_FILL, "RED": RED_FILL}

row = 4
for step in range(60):
    ts = base_time + timedelta(minutes=step)
    # Morning rush: traffic builds 8-9 AM
    rush_factor = 1.0 + 0.8 * math.sin(math.pi * step / 60)

    for int_id in intersections:
        queue = round(random.uniform(2, 20) * rush_factor, 1)
        vehicles = int(random.uniform(10, 50) * rush_factor)
        speed = round(random.uniform(3, 15) / rush_factor, 1)
        phase = random.choice(phases)
        phase_dur = round(random.uniform(10, 45), 1)
        action = random.randint(0, 3)
        waiting = round(queue * random.uniform(2, 5), 1)
        stopped = int(queue * random.uniform(0.3, 0.8))
        reward = round(-(waiting + 10 * stopped), 1)

        ws1.cell(row=row, column=1, value=ts.strftime("%Y-%m-%d %H:%M:%S"))
        ws1.cell(row=row, column=2, value=step + 1)
        ws1.cell(row=row, column=3, value=int_id)
        ws1.cell(row=row, column=4, value=queue)
        ws1.cell(row=row, column=5, value=vehicles)
        ws1.cell(row=row, column=6, value=speed)
        ws1.cell(row=row, column=7, value=phase)
        ws1.cell(row=row, column=8, value=phase_dur)
        ws1.cell(row=row, column=9, value=action)
        ws1.cell(row=row, column=10, value=reward)
        ws1.cell(row=row, column=11, value=waiting)
        ws1.cell(row=row, column=12, value=stopped)

        # Color signal phase cell
        if phase in phase_colors:
            ws1.cell(row=row, column=7).fill = phase_colors[phase]

        row += 1

style_data_rows(ws1, 4, row - 1, len(headers1))
auto_width(ws1, len(headers1))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Vehicle Detection Data (Camera)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Vehicle Detection")

ws2.merge_cells('A1:N1')
ws2.cell(row=1, column=1).value = "UrbanFlow — Vehicle Detection Data (Camera Feed)"
ws2.cell(row=1, column=1).font = TITLE_FONT

headers2 = [
    "Timestamp", "Frame #",
    "Total Vehicles", "Cars", "Buses", "Trucks", "Motorcycles",
    "North Lane", "South Lane", "East Lane", "West Lane",
    "Queue Length", "Avg Speed (m/s)", "Processing FPS"
]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=c, value=h)
style_header(ws2, 3, len(headers2))

row = 4
for frame in range(1, 101):
    ts = base_time + timedelta(seconds=frame * 0.2)
    total = random.randint(3, 25)
    cars = int(total * random.uniform(0.5, 0.7))
    buses = random.randint(0, 3)
    trucks = random.randint(0, 2)
    motos = total - cars - buses - trucks
    if motos < 0:
        motos = 0

    north = random.randint(0, total // 2 + 1)
    south = random.randint(0, total - north)
    east = random.randint(0, max(0, total - north - south))
    west = max(0, total - north - south - east)

    queue = random.randint(0, total // 2)
    speed = round(random.uniform(4, 15), 1)
    fps = round(random.uniform(1.0, 5.0), 1)

    ws2.cell(row=row, column=1, value=ts.strftime("%H:%M:%S.%f")[:-4])
    ws2.cell(row=row, column=2, value=frame)
    ws2.cell(row=row, column=3, value=total)
    ws2.cell(row=row, column=4, value=cars)
    ws2.cell(row=row, column=5, value=buses)
    ws2.cell(row=row, column=6, value=trucks)
    ws2.cell(row=row, column=7, value=motos)
    ws2.cell(row=row, column=8, value=north)
    ws2.cell(row=row, column=9, value=south)
    ws2.cell(row=row, column=10, value=east)
    ws2.cell(row=row, column=11, value=west)
    ws2.cell(row=row, column=12, value=queue)
    ws2.cell(row=row, column=13, value=speed)
    ws2.cell(row=row, column=14, value=fps)
    row += 1

style_data_rows(ws2, 4, row - 1, len(headers2))
auto_width(ws2, len(headers2))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Signal Detection Data
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Signal Detection")

ws3.merge_cells('A1:J1')
ws3.cell(row=1, column=1).value = "UrbanFlow — Traffic Signal Detection Data"
ws3.cell(row=1, column=1).font = TITLE_FONT

headers3 = [
    "Timestamp", "Frame #",
    "Signal State", "Confidence (%)",
    "Phase Duration (s)", "Cycle Time (s)",
    "Red Detected", "Yellow Detected", "Green Detected",
    "Circle Radius (px)"
]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=c, value=h)
style_header(ws3, 3, len(headers3))

row = 4
cycle_pos = 0
current_phase = "RED"
phase_timer = 0
cycle_time = 45.0

for frame in range(1, 101):
    ts = base_time + timedelta(seconds=frame * 0.2)
    phase_timer += 0.2

    # Simulate signal cycle: RED 20s -> GREEN 20s -> YELLOW 5s
    if current_phase == "RED" and phase_timer > 20:
        current_phase = "GREEN"
        phase_timer = 0
    elif current_phase == "GREEN" and phase_timer > 20:
        current_phase = "YELLOW"
        phase_timer = 0
    elif current_phase == "YELLOW" and phase_timer > 5:
        current_phase = "RED"
        phase_timer = 0

    conf = round(random.uniform(0.70, 0.98), 2)
    radius = random.randint(12, 20)

    ws3.cell(row=row, column=1, value=ts.strftime("%H:%M:%S.%f")[:-4])
    ws3.cell(row=row, column=2, value=frame)
    ws3.cell(row=row, column=3, value=current_phase)
    ws3.cell(row=row, column=4, value=round(conf * 100, 1))
    ws3.cell(row=row, column=5, value=round(phase_timer, 1))
    ws3.cell(row=row, column=6, value=cycle_time)
    ws3.cell(row=row, column=7, value="Yes" if current_phase == "RED" else "No")
    ws3.cell(row=row, column=8, value="Yes" if current_phase == "YELLOW" else "No")
    ws3.cell(row=row, column=9, value="Yes" if current_phase == "GREEN" else "No")
    ws3.cell(row=row, column=10, value=radius)

    if current_phase in phase_colors:
        ws3.cell(row=row, column=3).fill = phase_colors[current_phase]

    row += 1

style_data_rows(ws3, 4, row - 1, len(headers3))
auto_width(ws3, len(headers3))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: RL Agent Training Data
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("RL Agent Training")

ws4.merge_cells('A1:L1')
ws4.cell(row=1, column=1).value = "UrbanFlow — RL Agent Training Performance (100 Episodes)"
ws4.cell(row=1, column=1).font = TITLE_FONT

headers4 = [
    "Episode", "Total Steps", "Total Reward",
    "Avg Reward/Step", "Avg Queue", "Avg Speed (m/s)",
    "Avg Waiting Time (s)", "Vehicles Processed",
    "Epsilon (Exploration)", "Loss",
    "Training Time (s)", "Improvement (%)"
]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=c, value=h)
style_header(ws4, 3, len(headers4))

row = 4
initial_reward = -95.0
for episode in range(1, 101):
    steps = random.randint(800, 1200)
    # Reward improves over episodes (learning curve)
    improvement = 1.0 - math.exp(-episode / 30)
    total_reward = round(initial_reward * (1 - improvement * 0.7) * steps / 1000, 1)
    avg_reward = round(total_reward / steps, 3)
    avg_queue = round(20 * (1 - improvement * 0.6) + random.uniform(-2, 2), 1)
    avg_speed = round(5 + 10 * improvement * 0.8 + random.uniform(-1, 1), 1)
    avg_wait = round(50 * (1 - improvement * 0.5) + random.uniform(-5, 5), 1)
    vehicles = int(steps * random.uniform(0.3, 0.6))
    epsilon = round(max(0.01, 1.0 - episode * 0.01), 3)
    loss = round(5.0 * math.exp(-episode / 25) + random.uniform(0, 0.5), 4)
    train_time = round(random.uniform(15, 45), 1)
    impr_pct = round(improvement * 100, 1)

    ws4.cell(row=row, column=1, value=episode)
    ws4.cell(row=row, column=2, value=steps)
    ws4.cell(row=row, column=3, value=total_reward)
    ws4.cell(row=row, column=4, value=avg_reward)
    ws4.cell(row=row, column=5, value=avg_queue)
    ws4.cell(row=row, column=6, value=avg_speed)
    ws4.cell(row=row, column=7, value=avg_wait)
    ws4.cell(row=row, column=8, value=vehicles)
    ws4.cell(row=row, column=9, value=epsilon)
    ws4.cell(row=row, column=10, value=loss)
    ws4.cell(row=row, column=11, value=train_time)
    ws4.cell(row=row, column=12, value=impr_pct)
    row += 1

style_data_rows(ws4, 4, row - 1, len(headers4))
auto_width(ws4, len(headers4))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Lane-wise Analysis
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Lane Analysis")

ws5.merge_cells('A1:I1')
ws5.cell(row=1, column=1).value = "UrbanFlow — Lane-wise Traffic Analysis (Hourly Summary)"
ws5.cell(row=1, column=1).font = TITLE_FONT

headers5 = [
    "Time Period", "Lane",
    "Total Vehicles", "Avg Speed (m/s)", "Max Queue",
    "Avg Wait (s)", "Peak Count",
    "Throughput (veh/hr)", "Congestion Level"
]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=c, value=h)
style_header(ws5, 3, len(headers5))

lanes = ["NORTH", "SOUTH", "EAST", "WEST"]
time_periods = [
    "06:00-07:00", "07:00-08:00", "08:00-09:00", "09:00-10:00",
    "10:00-11:00", "11:00-12:00", "12:00-13:00", "13:00-14:00",
    "14:00-15:00", "15:00-16:00", "16:00-17:00", "17:00-18:00",
    "18:00-19:00", "19:00-20:00", "20:00-21:00", "21:00-22:00"
]

# Rush hours: 8-10 AM and 5-7 PM
rush_map = {
    "06:00-07:00": 0.4, "07:00-08:00": 0.7, "08:00-09:00": 1.0, "09:00-10:00": 0.9,
    "10:00-11:00": 0.5, "11:00-12:00": 0.5, "12:00-13:00": 0.6, "13:00-14:00": 0.6,
    "14:00-15:00": 0.5, "15:00-16:00": 0.6, "16:00-17:00": 0.8, "17:00-18:00": 1.0,
    "18:00-19:00": 0.9, "19:00-20:00": 0.6, "20:00-21:00": 0.4, "21:00-22:00": 0.3
}

congestion_fill = {
    "Low": GREEN_FILL,
    "Medium": YELLOW_FILL,
    "High": RED_FILL
}

row = 4
for period in time_periods:
    rush = rush_map[period]
    for lane in lanes:
        lane_bias = {"NORTH": 1.0, "SOUTH": 0.9, "EAST": 0.7, "WEST": 0.8}[lane]
        total_v = int(random.uniform(80, 400) * rush * lane_bias)
        avg_spd = round(random.uniform(5, 15) / max(rush, 0.5), 1)
        max_q = int(random.uniform(5, 25) * rush)
        avg_w = round(random.uniform(10, 60) * rush, 1)
        peak = int(total_v * random.uniform(0.08, 0.15))
        throughput = int(total_v * random.uniform(0.85, 1.0))
        cong = "High" if rush > 0.8 else ("Medium" if rush > 0.5 else "Low")

        ws5.cell(row=row, column=1, value=period)
        ws5.cell(row=row, column=2, value=lane)
        ws5.cell(row=row, column=3, value=total_v)
        ws5.cell(row=row, column=4, value=avg_spd)
        ws5.cell(row=row, column=5, value=max_q)
        ws5.cell(row=row, column=6, value=avg_w)
        ws5.cell(row=row, column=7, value=peak)
        ws5.cell(row=row, column=8, value=throughput)
        ws5.cell(row=row, column=9, value=cong)

        if cong in congestion_fill:
            ws5.cell(row=row, column=9).fill = congestion_fill[cong]

        row += 1

style_data_rows(ws5, 4, row - 1, len(headers5))
auto_width(ws5, len(headers5))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: System Performance Comparison
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Performance Comparison")

ws6.merge_cells('A1:G1')
ws6.cell(row=1, column=1).value = "UrbanFlow — AI vs Fixed Timer Performance Comparison"
ws6.cell(row=1, column=1).font = TITLE_FONT

headers6 = [
    "Metric", "Fixed Timer", "AI Adaptive",
    "Improvement", "Improvement (%)",
    "Unit", "Notes"
]
for c, h in enumerate(headers6, 1):
    ws6.cell(row=3, column=c, value=h)
style_header(ws6, 3, len(headers6))

comparisons = [
    ("Average Wait Time",       45.2, 28.1, "seconds",  "Reduced by AI optimization"),
    ("Average Queue Length",    12.5,  7.8, "vehicles", "Better phase allocation"),
    ("Average Speed",            6.8, 10.2, "m/s",      "Higher throughput"),
    ("Stopped Vehicles",        18.0, 10.5, "count",    "Fewer unnecessary stops"),
    ("Fuel Consumption Index",  100,   72,  "index",    "Baseline = 100"),
    ("CO2 Emissions Index",     100,   68,  "index",    "Baseline = 100"),
    ("Throughput (veh/hr)",     850, 1120,  "veh/hr",   "More vehicles processed"),
    ("Emergency Response Time",  4.5,  2.8, "minutes",  "Faster emergency clearance"),
    ("Peak Hour Congestion",    85,    58,  "%",        "Less congestion during rush"),
    ("Cycle Efficiency",        60,    82,  "%",        "Better green time usage"),
]

row = 4
for metric, fixed, ai, unit, note in comparisons:
    diff = round(ai - fixed, 1)
    pct = round(((ai - fixed) / fixed) * 100, 1)

    ws6.cell(row=row, column=1, value=metric)
    ws6.cell(row=row, column=2, value=fixed)
    ws6.cell(row=row, column=3, value=ai)
    ws6.cell(row=row, column=4, value=diff)
    ws6.cell(row=row, column=5, value=pct)
    ws6.cell(row=row, column=6, value=unit)
    ws6.cell(row=row, column=7, value=note)

    # Color improvement
    if ("Wait" in metric or "Queue" in metric or "Stopped" in metric or
            "Consumption" in metric or "Emission" in metric or "Congestion" in metric or
            "Response" in metric):
        # Lower is better for these
        cell_fill = GREEN_FILL if ai < fixed else RED_FILL
    else:
        # Higher is better
        cell_fill = GREEN_FILL if ai > fixed else RED_FILL
    ws6.cell(row=row, column=5).fill = cell_fill

    row += 1

style_data_rows(ws6, 4, row - 1, len(headers6))
auto_width(ws6, len(headers6))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7: HSV Color Ranges (Configuration)
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Detection Config")

ws7.merge_cells('A1:H1')
ws7.cell(row=1, column=1).value = "UrbanFlow — Detection Configuration Parameters"
ws7.cell(row=1, column=1).font = TITLE_FONT

# Signal detection HSV ranges
ws7.cell(row=3, column=1).value = "Signal Detection — HSV Color Ranges"
ws7.cell(row=3, column=1).font = SUBTITLE_FONT

headers7a = ["Color", "H Low", "H High", "S Low", "S High", "V Low", "V High", "Purpose"]
for c, h in enumerate(headers7a, 1):
    ws7.cell(row=4, column=c, value=h)
style_header(ws7, 4, len(headers7a))

hsv_data = [
    ("Red (Range 1)",   0,  10, 100, 255, 100, 255, "Red traffic signal"),
    ("Red (Range 2)", 160, 180, 100, 255, 100, 255, "Red signal (wraparound)"),
    ("Yellow",         15,  35, 100, 255, 100, 255, "Yellow traffic signal"),
    ("Green",          40,  90,  50, 255,  50, 255, "Green traffic signal"),
]

for i, (color, hl, hh, sl, sh, vl, vh, purpose) in enumerate(hsv_data):
    r = 5 + i
    ws7.cell(row=r, column=1, value=color)
    ws7.cell(row=r, column=2, value=hl)
    ws7.cell(row=r, column=3, value=hh)
    ws7.cell(row=r, column=4, value=sl)
    ws7.cell(row=r, column=5, value=sh)
    ws7.cell(row=r, column=6, value=vl)
    ws7.cell(row=r, column=7, value=vh)
    ws7.cell(row=r, column=8, value=purpose)

style_data_rows(ws7, 5, 8, len(headers7a))

# Vehicle detection config
ws7.cell(row=11, column=1).value = "Vehicle Detection — Parameters"
ws7.cell(row=11, column=1).font = SUBTITLE_FONT

headers7b = ["Parameter", "Value", "Unit", "Description"]
for c, h in enumerate(headers7b, 1):
    ws7.cell(row=12, column=c, value=h)
style_header(ws7, 12, len(headers7b))

veh_config = [
    ("MOG2 History",         200,    "frames",  "Background model memory"),
    ("MOG2 varThreshold",    25,     "-",       "Sensitivity (lower = more sensitive)"),
    ("Detect Shadows",       "No",   "-",       "Shadow artifact filtering"),
    ("Min Contour Area",     500,    "px²",     "Minimum blob size to count"),
    ("Max Contour Area",     "50%",  "frame",   "Maximum (filters full-frame noise)"),
    ("Aspect Ratio Min",     0.2,    "-",       "Vehicle shape lower bound"),
    ("Aspect Ratio Max",     4.0,    "-",       "Vehicle shape upper bound"),
    ("Stopped Threshold",    2.0,    "m/s",     "Speed below = queued"),
    ("Pixels Per Meter",     10,     "px/m",    "Calibration factor"),
    ("YOLO Model",           "yolov8n", "-",    "YOLOv8 Nano (6.5MB)"),
    ("YOLO Confidence",      0.5,    "-",       "Min detection confidence"),
    ("YOLO IOU Threshold",   0.45,   "-",       "Non-max suppression overlap"),
    ("Warmup Frames",        30,     "frames",  "Frames before reliable output"),
]

for i, (param, val, unit, desc) in enumerate(veh_config):
    r = 13 + i
    ws7.cell(row=r, column=1, value=param)
    ws7.cell(row=r, column=2, value=val)
    ws7.cell(row=r, column=3, value=unit)
    ws7.cell(row=r, column=4, value=desc)

style_data_rows(ws7, 13, 13 + len(veh_config) - 1, len(headers7b))
auto_width(ws7, max(len(headers7a), len(headers7b)))


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8: COCO Vehicle Classes
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Vehicle Classes")

ws8.merge_cells('A1:D1')
ws8.cell(row=1, column=1).value = "UrbanFlow — COCO Vehicle Class IDs (YOLO)"
ws8.cell(row=1, column=1).font = TITLE_FONT

headers8 = ["Class ID", "Vehicle Type", "Priority Weight", "Color Code"]
for c, h in enumerate(headers8, 1):
    ws8.cell(row=3, column=c, value=h)
style_header(ws8, 3, len(headers8))

classes = [
    (1, "Bicycle",    0.3, "#3B82F6"),
    (2, "Car",        1.0, "#22C55E"),
    (3, "Motorcycle", 0.5, "#F59E0B"),
    (5, "Bus",        2.5, "#EF4444"),
    (7, "Truck",      2.0, "#8B5CF6"),
]
for i, (cid, name, weight, color) in enumerate(classes):
    r = 4 + i
    ws8.cell(row=r, column=1, value=cid)
    ws8.cell(row=r, column=2, value=name)
    ws8.cell(row=r, column=3, value=weight)
    ws8.cell(row=r, column=4, value=color)

style_data_rows(ws8, 4, 4 + len(classes) - 1, len(headers8))
auto_width(ws8, len(headers8))


# ══════════════════════════════════════════════════════════════════════════════
# Add charts to key sheets
# ══════════════════════════════════════════════════════════════════════════════

# Chart: RL Training Reward Curve (Sheet 4)
chart1 = LineChart()
chart1.title = "RL Agent — Training Reward Curve"
chart1.x_axis.title = "Episode"
chart1.y_axis.title = "Total Reward"
chart1.style = 10
chart1.width = 25
chart1.height = 15

data_ref = Reference(ws4, min_col=3, min_row=3, max_row=103)
cats_ref = Reference(ws4, min_col=1, min_row=4, max_row=103)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.series[0].graphicalProperties.line.width = 20000
ws4.add_chart(chart1, "A106")

# Chart: Performance Comparison Bar Chart (Sheet 6)
chart2 = BarChart()
chart2.type = "col"
chart2.title = "AI vs Fixed Timer — Key Metrics"
chart2.style = 10
chart2.width = 25
chart2.height = 15

data_ref2 = Reference(ws6, min_col=2, max_col=3, min_row=3, max_row=13)
cats_ref2 = Reference(ws6, min_col=1, min_row=4, max_row=13)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
ws6.add_chart(chart2, "A16")


# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
output_path = "D:/Hackathon/UrbanFlow_System_Data.xlsx"
wb.save(output_path)
print(f"Excel file saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print("Done!")
